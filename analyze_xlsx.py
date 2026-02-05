"""
XLSX文件VALUE差异分析工具
比较多个xlsx文件中相同sheet、相同step的VALUE值差异、持续时长差异和趋势差异
优化版本：使用openpyxl read_only模式 + 多进程并行 + 内存缓存
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing


def load_config(config_path: str = "config.json") -> dict:
    """加载配置文件"""
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_all_xlsx_files(folder_path: str) -> list[str]:
    """获取文件夹中所有xlsx文件"""
    folder = Path(folder_path)
    return [str(f) for f in folder.glob("*.xlsx")]


def get_baseline_for_recipe(recipe: str, baseline_files: dict[str, str], folder_path: str) -> str | None:
    """根据Recipe后缀获取对应的基准文件路径"""
    for suffix, baseline_file in baseline_files.items():
        if recipe.endswith(suffix):
            baseline_path = Path(folder_path) / baseline_file
            if baseline_path.exists():
                return str(baseline_path)
    return None


def read_file_meta_only(file_path: str, meta_sheet: str, meta_columns: list[str]) -> dict:
    """只读取文件的元数据"""
    try:
        xl = pd.ExcelFile(file_path, engine='openpyxl')
        if meta_sheet not in xl.sheet_names:
            xl.close()
            return {}
        df = xl.parse(meta_sheet)
        xl.close()

        result = {}
        for col in meta_columns:
            if col in df.columns and len(df) > 0:
                value = df[col].iloc[0]
                result[col] = value if pd.notna(value) else ""
            else:
                result[col] = ""
        return result
    except Exception:
        return {}


def get_sheet_names_fast(file_path: str, specified_sheets: list[str]) -> list[str]:
    """获取要分析的sheet名称列表（使用openpyxl）"""
    wb = load_workbook(file_path, read_only=True)
    all_sheets = wb.sheetnames
    wb.close()

    if specified_sheets:
        return [s for s in specified_sheets if s in all_sheets]
    return all_sheets


def read_all_sheets_fast(file_path: str, sheets_to_read: list[str]) -> dict[str, pd.DataFrame]:
    """使用pandas一次性读取多个sheet"""
    result = {}
    try:
        xl = pd.ExcelFile(file_path, engine='openpyxl')
        available_sheets = set(xl.sheet_names)

        for sheet_name in sheets_to_read:
            if sheet_name not in available_sheets:
                continue
            try:
                result[sheet_name] = xl.parse(sheet_name)
            except Exception:
                result[sheet_name] = pd.DataFrame()

        xl.close()
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")

    return result


def normalize_step_value(step_val) -> str | None:
    """标准化STEP列的值，将数字或文本转为整数字符串"""
    if pd.isna(step_val):
        return None
    if isinstance(step_val, (int, float)):
        return str(int(step_val))
    try:
        return str(int(float(str(step_val).strip())))
    except (ValueError, TypeError):
        return str(step_val).strip()


def parse_sheet_data_vectorized(df: pd.DataFrame, step_col: str, value_col: str) -> dict:
    """向量化解析sheet数据，返回 {step: {'values': [values], 'count': 行数}} 字典"""
    if df.empty or step_col not in df.columns or value_col not in df.columns:
        return {}

    df = df.copy()

    # 向量化标准化STEP列
    def normalize_step(val):
        if pd.isna(val):
            return None
        if isinstance(val, (int, float)):
            return str(int(val))
        try:
            return str(int(float(str(val).strip())))
        except:
            return str(val).strip()

    df['_step_norm'] = df[step_col].apply(normalize_step)
    df['_value_num'] = pd.to_numeric(df[value_col], errors='coerce')

    # 过滤有效数据
    valid = df.dropna(subset=['_step_norm', '_value_num'])

    result = {}
    for step, group in valid.groupby('_step_norm'):
        values = group['_value_num'].tolist()
        result[step] = {'values': values, 'count': len(values)}

    return result


def read_meta_info(df: pd.DataFrame, meta_columns: list[str]) -> dict:
    """从DataFrame读取元数据"""
    result = {}
    for col in meta_columns:
        if col in df.columns and len(df) > 0:
            value = df[col].iloc[0]
            result[col] = value if pd.notna(value) else ""
        else:
            result[col] = ""
    return result


def load_file_data(file_path: str, sheets: list[str], step_col: str, value_col: str,
                   meta_sheet: str = None, meta_columns: list[str] = None) -> tuple[dict, dict]:
    """加载单个文件的所有sheet数据和元数据"""
    all_sheets = list(sheets)
    if meta_sheet and meta_sheet not in all_sheets:
        all_sheets.append(meta_sheet)

    sheets_df = read_all_sheets_fast(file_path, all_sheets)
    file_data = {}

    for sheet_name, df in sheets_df.items():
        if sheet_name != meta_sheet:
            file_data[sheet_name] = parse_sheet_data_vectorized(df, step_col, value_col)

    # 读取元数据
    meta_info = {}
    if meta_sheet and meta_columns:
        meta_df = sheets_df.get(meta_sheet, pd.DataFrame())
        meta_info = read_meta_info(meta_df, meta_columns)

    return file_data, meta_info


def load_file_wrapper(args: tuple) -> tuple[str, dict, dict]:
    """多进程包装函数"""
    file_path, sheets, step_col, value_col, meta_sheet, meta_columns = args
    try:
        data, meta = load_file_data(file_path, sheets, step_col, value_col, meta_sheet, meta_columns)
        return (file_path, data, meta)
    except Exception as e:
        print(f"加载文件失败 {file_path}: {e}")
        return (file_path, {}, {})


def calculate_value_diff(baseline_values: list[float], compare_values: list[float]) -> tuple[float, float, float]:
    """计算VALUE差异百分比，返回 (基准均值, 比较均值, 差异百分比)"""
    baseline_avg = sum(baseline_values) / len(baseline_values)
    compare_avg = sum(compare_values) / len(compare_values)

    if baseline_avg == 0:
        diff_percent = 100.0 if compare_avg != 0 else 0.0
    else:
        diff_percent = abs(compare_avg - baseline_avg) / abs(baseline_avg) * 100

    return baseline_avg, compare_avg, diff_percent


def calculate_correlation(baseline_values: list[float], compare_values: list[float]) -> float:
    """
    计算Pearson相关系数
    如果长度不同，取较短长度进行对比
    返回相关系数 [-1, 1]，1表示完全正相关，-1表示完全负相关，0表示无相关
    """
    min_len = min(len(baseline_values), len(compare_values))
    if min_len < 3:
        return 1.0  # 数据点太少，无法判断

    arr1 = np.array(baseline_values[:min_len])
    arr2 = np.array(compare_values[:min_len])

    # 检查是否全部相同（标准差为0）
    if np.std(arr1) == 0 or np.std(arr2) == 0:
        return 1.0 if np.mean(arr1) == np.mean(arr2) else 0.0

    corr = np.corrcoef(arr1, arr2)[0, 1]
    return corr if not np.isnan(corr) else 1.0


def calculate_segment_trend(values: list[float], num_segments: int) -> list[int]:
    """
    计算分段趋势，返回每段的变化方向列表
    1: 上升, -1: 下降, 0: 平稳
    """
    if len(values) < 2:
        return [0]

    segment_size = max(1, len(values) // num_segments)
    trends = []

    for i in range(num_segments):
        start = i * segment_size
        end = start + segment_size if i < num_segments - 1 else len(values)

        if start >= len(values):
            break

        segment = values[start:end]
        if len(segment) < 2:
            trends.append(0)
            continue

        first_half_avg = np.mean(segment[:len(segment)//2]) if len(segment) >= 2 else segment[0]
        second_half_avg = np.mean(segment[len(segment)//2:]) if len(segment) >= 2 else segment[-1]
        diff = second_half_avg - first_half_avg

        if abs(diff) < 1e-6:
            trends.append(0)
        elif diff > 0:
            trends.append(1)
        else:
            trends.append(-1)

    return trends


def compare_segment_trends(baseline_values: list[float], compare_values: list[float], num_segments: int) -> tuple[float, list[int], list[int]]:
    """
    比较分段趋势，返回 (一致性百分比, 基准趋势, 比较趋势)
    """
    baseline_trends = calculate_segment_trend(baseline_values, num_segments)
    compare_trends = calculate_segment_trend(compare_values, num_segments)

    min_len = min(len(baseline_trends), len(compare_trends))
    if min_len == 0:
        return 100.0, baseline_trends, compare_trends

    matches = sum(1 for i in range(min_len) if baseline_trends[i] == compare_trends[i])
    consistency = (matches / min_len) * 100

    return consistency, baseline_trends, compare_trends


def calculate_diff_signs(values: list[float]) -> list[int]:
    """
    计算差分符号序列
    1: 上升, -1: 下降, 0: 不变
    """
    if len(values) < 2:
        return []

    signs = []
    for i in range(1, len(values)):
        diff = values[i] - values[i-1]
        if abs(diff) < 1e-6:
            signs.append(0)
        elif diff > 0:
            signs.append(1)
        else:
            signs.append(-1)

    return signs


def compare_diff_signs(baseline_values: list[float], compare_values: list[float]) -> tuple[float, int, int]:
    """
    比较差分符号序列，返回 (一致性百分比, 基准序列长度, 比较序列长度)
    """
    baseline_signs = calculate_diff_signs(baseline_values)
    compare_signs = calculate_diff_signs(compare_values)

    min_len = min(len(baseline_signs), len(compare_signs))
    if min_len == 0:
        return 100.0, len(baseline_signs), len(compare_signs)

    matches = sum(1 for i in range(min_len) if baseline_signs[i] == compare_signs[i])
    consistency = (matches / min_len) * 100

    return consistency, len(baseline_signs), len(compare_signs)


def trend_to_str(trends: list[int]) -> str:
    """将趋势列表转为可读字符串"""
    symbols = {1: "↑", -1: "↓", 0: "→"}
    return "".join(symbols.get(t, "?") for t in trends)


def analyze_files(config: dict) -> tuple[dict, list[str]]:
    """分析所有文件，返回各类差异结果（根据Recipe后缀使用不同基准文件）"""
    folder_path = config["folder_path"]
    baseline_files = config.get("baseline_files", {})
    recipe_column = config.get("recipe_column", "recipe")

    # 阈值配置
    value_threshold = config["threshold_percent"]
    duration_threshold = config.get("duration_threshold_rows", 1)
    correlation_threshold = config.get("correlation_threshold", 0.8)
    segment_threshold = config.get("segment_trend_threshold", 80)
    diff_sign_threshold = config.get("diff_sign_threshold", 70)
    num_segments = config.get("num_segments", 4)

    step_col = config["step_column"]
    value_col = config["value_column"]
    meta_sheet = config.get("meta_sheet", "")
    meta_columns = config.get("meta_columns", [])

    results = {
        "value": [],
        "duration": [],
        "correlation": [],
        "segment": [],
        "diff_sign": []
    }

    # 获取所有文件
    all_files = get_all_xlsx_files(folder_path)
    baseline_file_names = set(baseline_files.values())
    compare_files = [f for f in all_files if Path(f).name not in baseline_file_names]

    print(f"基准文件配置: {baseline_files}")
    print(f"待比较文件数: {len(compare_files)}")
    print(f"VALUE差异阈值: {value_threshold}%")
    print(f"时长差异阈值: {duration_threshold}行")
    print(f"相关系数阈值: {correlation_threshold}")
    print(f"分段趋势一致性阈值: {segment_threshold}%")
    print(f"差分符号一致性阈值: {diff_sign_threshold}%")
    print("-" * 60)

    # 第一阶段：读取所有文件的元数据，按Recipe分组
    print("阶段1: 读取文件元数据并分组...")
    file_groups = {}  # {baseline_path: [compare_files]}

    for file_path in compare_files:
        meta = read_file_meta_only(file_path, meta_sheet, meta_columns)
        recipe = str(meta.get(recipe_column, ""))
        baseline_path = get_baseline_for_recipe(recipe, baseline_files, folder_path)

        if baseline_path:
            if baseline_path not in file_groups:
                file_groups[baseline_path] = []
            file_groups[baseline_path].append(file_path)
        else:
            print(f"  警告: {Path(file_path).name} 的Recipe '{recipe}' 无匹配基准文件，跳过")

    for baseline_path, files in file_groups.items():
        print(f"  {Path(baseline_path).name}: {len(files)} 个文件")

    # 第二阶段：按组加载数据并比较
    print("\n阶段2: 并行加载文件数据并比较...")

    for baseline_file, group_files in file_groups.items():
        if not group_files:
            continue

        print(f"\n处理基准文件: {Path(baseline_file).name}")

        # 获取要分析的sheet
        sheets = get_sheet_names_fast(baseline_file, config.get("sheets", []))

        # 加载该组的所有文件
        all_group_files = [baseline_file] + group_files
        num_workers = min(multiprocessing.cpu_count(), len(all_group_files))

        tasks = [(f, sheets, step_col, value_col, meta_sheet, meta_columns) for f in all_group_files]
        files_data = {}
        files_meta = {}

        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            futures = {executor.submit(load_file_wrapper, task): task[0] for task in tasks}
            for future in as_completed(futures):
                file_path, data, meta = future.result()
                files_data[file_path] = data
                files_meta[file_path] = meta

        print(f"  加载完成，开始比较 {len(group_files)} 个文件...")

        # 比较分析
        baseline_data_all = files_data.get(baseline_file, {})

        for sheet in sheets:
            baseline_data = baseline_data_all.get(sheet, {})
            if not baseline_data:
                continue

            for compare_file in group_files:
                compare_name = Path(compare_file).name
                compare_meta = files_meta.get(compare_file, {})
                compare_data_all = files_data.get(compare_file, {})
                compare_data = compare_data_all.get(sheet, {})
                if not compare_data:
                    continue

                for step, baseline_info in baseline_data.items():
                    if step not in compare_data:
                        continue

                    compare_info = compare_data[step]
                    baseline_values = baseline_info['values']
                    compare_values = compare_info['values']
                    baseline_count = baseline_info['count']
                    compare_count = compare_info['count']

                    # 基础结果信息（包含元数据）
                    base_result = {"比较文件": compare_name, **compare_meta, "Sheet名": sheet, "Step": step}

                    # 1. 检查VALUE差异
                    baseline_avg, compare_avg, diff_percent = calculate_value_diff(baseline_values, compare_values)
                    if diff_percent > value_threshold:
                        results["value"].append({
                            **base_result,
                            "基准均值": round(baseline_avg, 4), "比较均值": round(compare_avg, 4),
                            "差异百分比": round(diff_percent, 2)
                        })

                    # 2. 检查时长(行数)差异
                    duration_diff = abs(compare_count - baseline_count)
                    if duration_diff > duration_threshold:
                        results["duration"].append({
                            **base_result,
                            "基准行数": baseline_count, "比较行数": compare_count,
                            "差异行数": duration_diff
                        })

                    # 3. 检查相关系数（趋势方向）
                    corr = calculate_correlation(baseline_values, compare_values)
                    if corr < correlation_threshold:
                        results["correlation"].append({
                            **base_result,
                            "相关系数": round(corr, 4)
                        })

                    # 4. 检查分段趋势
                    seg_consistency, baseline_trends, compare_trends = compare_segment_trends(
                        baseline_values, compare_values, num_segments
                    )
                    if seg_consistency < segment_threshold:
                        results["segment"].append({
                            **base_result,
                            "一致性": round(seg_consistency, 1),
                            "基准趋势": trend_to_str(baseline_trends),
                            "比较趋势": trend_to_str(compare_trends)
                        })

                    # 5. 检查差分符号序列
                    diff_consistency, _, _ = compare_diff_signs(baseline_values, compare_values)
                    if diff_consistency < diff_sign_threshold:
                        results["diff_sign"].append({
                            **base_result,
                            "一致性": round(diff_consistency, 1)
                        })

    return results, meta_columns


def save_results(results: dict, meta_columns: list[str], output_file: str):
    """保存结果到CSV文件"""
    all_rows = []

    def build_row(r: dict, diff_type: str, detail: str) -> dict:
        row = {"比较文件": r["比较文件"]}
        for col in meta_columns:
            row[col] = r.get(col, "")
        row["Sheet名"] = r["Sheet名"]
        row["Step"] = r["Step"]
        row["差异类型"] = diff_type
        row["详情"] = detail
        return row

    for r in results["value"]:
        all_rows.append(build_row(r, "VALUE均值",
            f"基准:{r['基准均值']} 比较:{r['比较均值']} 差异:{r['差异百分比']}%"))

    for r in results["duration"]:
        all_rows.append(build_row(r, "时长",
            f"基准:{r['基准行数']}行 比较:{r['比较行数']}行 差异:{r['差异行数']}行"))

    for r in results["correlation"]:
        all_rows.append(build_row(r, "相关系数", f"相关系数:{r['相关系数']}"))

    for r in results["segment"]:
        all_rows.append(build_row(r, "分段趋势",
            f"一致性:{r['一致性']}% 基准:{r['基准趋势']} 比较:{r['比较趋势']}"))

    for r in results["diff_sign"]:
        all_rows.append(build_row(r, "差分符号", f"一致性:{r['一致性']}%"))

    if all_rows:
        df = pd.DataFrame(all_rows)
        df.to_csv(output_file, index=False, encoding="utf-8-sig")
        print(f"\n结果已保存到: {output_file}")
    else:
        print("\n未发现超出阈值的差异")


def main():
    config = load_config()
    results, meta_columns = analyze_files(config)

    print(f"\n===== 分析结果汇总 =====")
    print(f"VALUE均值差异: {len(results['value'])}条")
    print(f"时长差异: {len(results['duration'])}条")
    print(f"相关系数异常: {len(results['correlation'])}条")
    print(f"分段趋势差异: {len(results['segment'])}条")
    print(f"差分符号差异: {len(results['diff_sign'])}条")

    total = sum(len(v) for v in results.values())
    print(f"\n共发现 {total} 条差异记录")

    save_results(results, meta_columns, config["output_file"])


if __name__ == "__main__":
    main()
