"""
XLSX文件特征提取工具
从多个xlsx文件中提取参数特征，生成汇总表
优化版本：使用openpyxl read_only模式 + 多进程并行 + 向量化操作
"""

import json
import re
import numpy as np
import pandas as pd
from pathlib import Path
from scipy.optimize import curve_fit
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing


def load_config(config_path: str = "config_feature.json") -> dict:
    """加载配置文件"""
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_step_config(steps_config: list[str]) -> list[str]:
    """
    解析STEP配置，支持单个STEP和范围
    配置格式：STEP18 表示值18，STEP20-23 表示值20,21,22,23
    例如：["STEP18", "STEP20-23"] -> ["18", "20", "21", "22", "23"]
    """
    result = []
    for item in steps_config:
        if "-" in item:
            match = re.match(r"(\D*)(\d+)-(\d+)", item)
            if match:
                start = int(match.group(2))
                end = int(match.group(3))
                for i in range(start, end + 1):
                    result.append(str(i))
        else:
            match = re.match(r"(\D*)(\d+)$", item)
            if match:
                result.append(match.group(2))
            else:
                result.append(item)
    return result


def get_step_groups(steps_config: list[str]) -> dict[str, list[str]]:
    """
    获取STEP分组，用于合并分析
    返回 {组名: [STEP值列表]}
    """
    groups = {}
    for item in steps_config:
        if "-" in item:
            match = re.match(r"(\D*)(\d+)-(\d+)", item)
            if match:
                start = int(match.group(2))
                end = int(match.group(3))
                steps = [str(i) for i in range(start, end + 1)]
                groups[item] = steps
        else:
            match = re.match(r"(\D*)(\d+)$", item)
            if match:
                num = match.group(2)
                groups[item] = [num]
            else:
                groups[item] = [item]
    return groups


def read_all_sheets_fast(file_path: str, sheets_to_read: list[str]) -> dict[str, pd.DataFrame]:
    """
    使用openpyxl read_only模式快速读取多个sheet
    一次打开文件，读取所有需要的sheet
    """
    result = {}
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = set(wb.sheetnames)

        for sheet_name in sheets_to_read:
            if sheet_name not in available_sheets:
                continue

            ws = wb[sheet_name]
            data = list(ws.iter_rows(values_only=True))

            if len(data) < 2:
                result[sheet_name] = pd.DataFrame()
                continue

            headers = data[0]
            rows = data[1:]
            result[sheet_name] = pd.DataFrame(rows, columns=headers)

        wb.close()
    except Exception as e:
        pass

    return result


def extract_values_vectorized(df: pd.DataFrame, step_col: str, value_col: str, target_steps: set[str]) -> list[float]:
    """使用向量化操作提取目标STEP的VALUE数据"""
    if df.empty or step_col not in df.columns or value_col not in df.columns:
        return []

    df = df.copy()

    # 向量化标准化STEP列
    def normalize_step(val):
        if pd.isna(val):
            return ""
        if isinstance(val, (int, float)):
            return str(int(val))
        try:
            return str(int(float(str(val).strip())))
        except:
            return str(val).strip()

    df['_step_norm'] = df[step_col].apply(normalize_step)

    # 向量化筛选和转换
    mask = df['_step_norm'].isin(target_steps)
    filtered = df.loc[mask, value_col]

    values = pd.to_numeric(filtered, errors='coerce').dropna().tolist()
    return values


def read_meta_info_from_df(df: pd.DataFrame, meta_columns: list[str]) -> dict:
    """从DataFrame读取元数据"""
    result = {}
    for col in meta_columns:
        if col in df.columns and len(df) > 0:
            value = df[col].iloc[0]
            result[col] = value if pd.notna(value) else ""
        else:
            result[col] = ""
    return result


def is_stable(values: list[float], threshold: float) -> bool:
    """判断数据是否稳定（值相同或浮动在阈值内）"""
    if len(values) < 2:
        return True
    return (max(values) - min(values)) <= threshold


def exp_func(x, a, b, c):
    """指数函数 y = a(1 - e^(-bx)) + c"""
    return a * (1 - np.exp(-b * x)) + c


def fit_temperature_curve(values: list[float]) -> tuple[float | None, float | None]:
    """
    拟合温度曲线 y = a(1 - e^(-bx)) + c
    返回 (b值, 均值) - 如果拟合成功返回b，否则返回None
    """
    if len(values) < 4:
        return None, np.mean(values) if values else None

    x = np.arange(len(values))
    y = np.array(values)

    try:
        # 初始猜测
        a_init = max(y) - min(y)
        c_init = min(y)
        b_init = 0.1

        popt, _ = curve_fit(
            exp_func, x, y,
            p0=[a_init, b_init, c_init],
            bounds=([0, 0.001, -np.inf], [np.inf, 10, np.inf]),
            maxfev=5000
        )
        return popt[1], np.mean(values)  # 返回b值和均值
    except Exception:
        return None, np.mean(values)


def count_fluctuations(values: list[float]) -> int:
    """统计0/1数据的波动次数（状态切换次数）"""
    if len(values) < 2:
        return 0

    count = 0
    for i in range(1, len(values)):
        if values[i] != values[i-1]:
            count += 1
    return count


def count_stair_steps(values: list[float], threshold: float) -> int:
    """
    统计阶梯型数据的台阶数量
    通过检测差分超过阈值的跳变点来识别台阶
    台阶数量 = 跳变次数 + 1
    """
    if len(values) < 2:
        return 1 if values else 0

    jump_count = 0
    for i in range(1, len(values)):
        if abs(values[i] - values[i-1]) > threshold:
            jump_count += 1

    return jump_count + 1


def find_stable_point(values: list[float], threshold: float) -> int:
    """
    找到数据开始稳定的位置
    返回稳定开始的索引，如果没找到返回-1
    """
    if len(values) < 5:
        return -1

    window_size = 5
    for i in range(len(values) - window_size + 1):
        window = values[i:i + window_size]
        if (max(window) - min(window)) <= threshold:
            return i

    return -1


def analyze_trend(values: list[float], threshold: float) -> dict:
    """
    分析数据趋势
    返回特征字典
    """
    if len(values) < 2:
        return {
            "slope": None,
            "fluctuations_before_stable": 0,
            "stable_max": values[0] if values else None,
            "overall_max": values[0] if values else None,
            "overall_min": values[0] if values else None,
            "trend_type": "insufficient_data"
        }

    overall_max = max(values)
    overall_min = min(values)

    # 找稳定点
    stable_idx = find_stable_point(values, threshold)

    if stable_idx == -1:
        # 没有稳定段，全部是波动
        fluctuations = count_fluctuations(values)
        return {
            "slope": None,
            "fluctuations_before_stable": fluctuations,
            "stable_max": None,
            "overall_max": overall_max,
            "overall_min": overall_min,
            "trend_type": "no_stable"
        }

    # 有稳定段
    before_stable = values[:stable_idx] if stable_idx > 0 else []
    stable_part = values[stable_idx:]
    stable_max = max(stable_part) if stable_part else None

    # 判断稳定前是上升、下降还是波动
    if len(before_stable) >= 2:
        first_val = before_stable[0]
        last_val = before_stable[-1]
        diff = last_val - first_val

        # 计算单调性
        increases = sum(1 for i in range(1, len(before_stable)) if before_stable[i] > before_stable[i-1])
        decreases = sum(1 for i in range(1, len(before_stable)) if before_stable[i] < before_stable[i-1])
        total_changes = len(before_stable) - 1

        # 判断是否单调
        monotonic_ratio = max(increases, decreases) / total_changes if total_changes > 0 else 0

        if monotonic_ratio >= 0.7:
            # 单调上升或下降
            x = np.arange(len(before_stable))
            slope = np.polyfit(x, before_stable, 1)[0]
            trend_type = "rising" if slope > 0 else "falling"
            return {
                "slope": slope,
                "fluctuations_before_stable": 0,
                "stable_max": stable_max,
                "overall_max": overall_max,
                "overall_min": overall_min,
                "trend_type": trend_type
            }
        else:
            # 波动
            fluctuations = count_fluctuations(before_stable)
            return {
                "slope": None,
                "fluctuations_before_stable": fluctuations,
                "stable_max": stable_max,
                "overall_max": overall_max,
                "overall_min": overall_min,
                "trend_type": "fluctuating"
            }
    else:
        # 几乎立即稳定
        return {
            "slope": None,
            "fluctuations_before_stable": 0,
            "stable_max": stable_max,
            "overall_max": overall_max,
            "overall_min": overall_min,
            "trend_type": "immediate_stable"
        }


def process_file(file_path: str, config: dict) -> dict:
    """处理单个文件，返回特征字典（优化版：一次读取所有sheet）"""
    result = {}

    # 收集所有需要读取的sheet
    all_sheets = [config["meta_sheet"]]
    all_sheets.extend(config.get("temperature_sheets", []))
    all_sheets.extend(config.get("binary_sheets", []))
    all_sheets.extend(config.get("other_sheets", []))
    all_sheets.extend(config.get("stair_sheets", []))

    # 一次性读取所有sheet
    sheets_data = read_all_sheets_fast(file_path, all_sheets)

    # 读取元数据
    meta_df = sheets_data.get(config["meta_sheet"], pd.DataFrame())
    meta = read_meta_info_from_df(meta_df, config["meta_columns"])
    result.update(meta)

    step_groups = get_step_groups(config["steps_to_analyze"])
    threshold = config["stable_threshold"]
    step_col = config["step_column"]
    value_col = config["value_column"]

    # 处理温度类sheet
    for sheet in config.get("temperature_sheets", []):
        df = sheets_data.get(sheet, pd.DataFrame())
        for group_name, steps in step_groups.items():
            col_prefix = f"{sheet}_{group_name}"
            target_steps = set(steps)
            values = extract_values_vectorized(df, step_col, value_col, target_steps)

            if not values:
                result[f"{col_prefix}_rate"] = None
                result[f"{col_prefix}_mean"] = None
                continue

            if is_stable(values, threshold):
                result[f"{col_prefix}_rate"] = None
                result[f"{col_prefix}_mean"] = np.mean(values)
            else:
                b_val, mean_val = fit_temperature_curve(values)
                result[f"{col_prefix}_rate"] = round(b_val, 6) if b_val else None
                result[f"{col_prefix}_mean"] = round(mean_val, 4) if mean_val else None

    # 处理0/1类sheet
    for sheet in config.get("binary_sheets", []):
        df = sheets_data.get(sheet, pd.DataFrame())
        for group_name, steps in step_groups.items():
            col_prefix = f"{sheet}_{group_name}"
            target_steps = set(steps)
            values = extract_values_vectorized(df, step_col, value_col, target_steps)
            fluctuations = count_fluctuations(values) if values else 0
            result[f"{col_prefix}_fluctuations"] = fluctuations

    # 处理其他类sheet
    for sheet in config.get("other_sheets", []):
        df = sheets_data.get(sheet, pd.DataFrame())
        for group_name, steps in step_groups.items():
            col_prefix = f"{sheet}_{group_name}"
            target_steps = set(steps)
            values = extract_values_vectorized(df, step_col, value_col, target_steps)

            if not values:
                result[f"{col_prefix}_slope"] = None
                result[f"{col_prefix}_fluctuations"] = None
                result[f"{col_prefix}_stable_max"] = None
                result[f"{col_prefix}_max"] = None
                result[f"{col_prefix}_min"] = None
                continue

            trend = analyze_trend(values, threshold)

            if trend["trend_type"] in ["rising", "falling"]:
                result[f"{col_prefix}_slope"] = round(trend["slope"], 6) if trend["slope"] else None
                result[f"{col_prefix}_fluctuations"] = None
            else:
                # 波动较大时，也计算整体斜率
                x = np.arange(len(values))
                slope = np.polyfit(x, values, 1)[0]
                result[f"{col_prefix}_slope"] = round(slope, 6)
                result[f"{col_prefix}_fluctuations"] = trend["fluctuations_before_stable"]

            result[f"{col_prefix}_stable_max"] = round(trend["stable_max"], 4) if trend["stable_max"] else None
            result[f"{col_prefix}_max"] = round(trend["overall_max"], 4) if trend["overall_max"] else None
            result[f"{col_prefix}_min"] = round(trend["overall_min"], 4) if trend["overall_min"] else None

    # 处理阶梯型sheet
    stair_threshold = config.get("stair_change_threshold", 0.5)
    for sheet in config.get("stair_sheets", []):
        df = sheets_data.get(sheet, pd.DataFrame())
        for group_name, steps in step_groups.items():
            col_prefix = f"{sheet}_{group_name}"
            target_steps = set(steps)
            values = extract_values_vectorized(df, step_col, value_col, target_steps)

            if not values:
                result[f"{col_prefix}_stair_count"] = None
                result[f"{col_prefix}_max"] = None
                result[f"{col_prefix}_min"] = None
                continue

            result[f"{col_prefix}_stair_count"] = count_stair_steps(values, stair_threshold)
            result[f"{col_prefix}_max"] = round(max(values), 4)
            result[f"{col_prefix}_min"] = round(min(values), 4)

    return result


def process_file_wrapper(args: tuple) -> dict:
    """多进程包装函数"""
    file_path, config = args
    try:
        result = process_file(file_path, config)
        result["源文件"] = Path(file_path).name
        return result
    except Exception as e:
        print(f"处理文件失败 {file_path}: {e}")
        return {"源文件": Path(file_path).name}


def main():
    config = load_config()
    folder = Path(config["folder_path"])
    xlsx_files = list(folder.glob("*.xlsx"))

    print(f"找到 {len(xlsx_files)} 个xlsx文件")
    print(f"STEP配置: {config['steps_to_analyze']}")
    print(f"温度类sheet: {config.get('temperature_sheets', [])}")
    print(f"0/1类sheet: {config.get('binary_sheets', [])}")
    print(f"其他类sheet: {config.get('other_sheets', [])}")
    print(f"阶梯类sheet: {config.get('stair_sheets', [])}")

    # 使用CPU核心数的进程池
    num_workers = min(multiprocessing.cpu_count(), len(xlsx_files))
    print(f"使用 {num_workers} 个并行进程")
    print("-" * 60)

    all_results = []
    tasks = [(str(f), config) for f in xlsx_files]

    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(process_file_wrapper, task): task[0] for task in tasks}
        completed = 0

        for future in as_completed(futures):
            completed += 1
            result = future.result()
            all_results.append(result)

            if completed % 50 == 0 or completed == len(xlsx_files):
                print(f"进度: {completed}/{len(xlsx_files)} ({completed*100//len(xlsx_files)}%)")

    if all_results:
        df = pd.DataFrame(all_results)
        meta_cols = ["源文件"] + config["meta_columns"]
        other_cols = [c for c in df.columns if c not in meta_cols]
        df = df[meta_cols + other_cols]

        df.to_excel(config["output_file"], index=False)
        print(f"\n结果已保存到: {config['output_file']}")
        print(f"共处理 {len(all_results)} 个文件，生成 {len(df.columns)} 列特征")
    else:
        print("\n未找到可处理的文件")


if __name__ == "__main__":
    main()
